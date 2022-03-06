import openpyxl

wb = openpyxl.load_workbook(excel_filename, data_only=True)
# print("シート名一覧："wb.sheetnames)
assert target_sheet_name in wb.sheetnames
sheet = wb[target_sheet_name]
dataset = []
for i in range(8,28):
    dataset.append([
        sheet.cell(row=i, column=k).value for k in [2,3,4,5]]
    )
